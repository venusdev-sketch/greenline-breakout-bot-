# format_glb_telegram.py
# Creates formatted Excel then sends Telegram message + optional attachment.
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os, sys, json, requests, traceback

CSV_FILE = "glb_signals.csv"
OUT_PREFIX = "GLB"
CFG_FILE = "notify_config.json"
TELE_API = "https://api.telegram.org"

def make_excel(df, out_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Breakouts"
    df_break = df[df.get('signal') == True].copy() if 'signal' in df.columns else df.copy()

    for r in dataframe_to_rows(df_break, index=False, header=True):
        ws1.append(r)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    if ws1.max_row >= 1:
        for cell in ws1[1]:
            try:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            except:
                pass

    ws1.auto_filter.ref = ws1.dimensions if ws1.max_row > 1 else "A1"

    green_fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            try:
                cell.fill = green_fill
            except:
                pass

    ws2 = wb.create_sheet("All Signals (Grouped)")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws2.append(r)

    if ws2.max_row >= 1:
        for cell in ws2[1]:
            try:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            except:
                pass

    ws2.auto_filter.ref = ws2.dimensions

    if 'signal' in df.columns:
        green_fill2 = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        red_fill   = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        sig_col = list(df.columns).index("signal") + 1
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            try:
                val = row[sig_col - 1].value
                fill = green_fill2 if val == True else red_fill
                for cell in row:
                    cell.fill = fill
            except:
                pass

    # Auto width
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            try:
                col_letter = col[0].column_letter
            except:
                continue
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            try:
                ws.column_dimensions[col_letter].width = max_len + 2
            except:
                pass

    wb.save(out_path)

def load_cfg(path):
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def send_telegram_message(bot_token, chat_id, text):
    url = f"{TELE_API}/bot{bot_token}/sendMessage"
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "HTML"}
    r = requests.post(url, data=payload, timeout=20)
    return r.ok, r.text

def send_telegram_file(bot_token, chat_id, file_path, caption=None):
    url = f"{TELE_API}/bot{bot_token}/sendDocument"
    with open(file_path, "rb") as f:
        files = {"document": (os.path.basename(file_path), f)}
        data = {"chat_id": chat_id}
        if caption:
            data["caption"] = caption
        r = requests.post(url, data=data, files=files, timeout=60)
    return r.ok, r.text

def main():
    try:
        print("INFO: format_glb_telegram.py starting")
        if not os.path.exists(CSV_FILE):
            print(f"ERROR: {CSV_FILE} not found.")
            return 2
        df = pd.read_csv(CSV_FILE)
        now = datetime.now().strftime("%Y-%m-%d_%H%M")
        out_name = f"{OUT_PREFIX}_{now}.xlsx"
        make_excel(df, out_name)
        print("Wrote", out_name)

        cfg = None
        try:
            cfg = load_cfg(CFG_FILE)
        except Exception as e:
            print("ERROR reading config:", e)
            traceback.print_exc()
            cfg = None

        if not cfg or "telegram" not in cfg:
            print("No notify_config.json/telegram config found — finishing without sending.")
            return 0

        tcfg = cfg["telegram"]
        bot_token = tcfg.get("bot_token")
        chat_id = tcfg.get("chat_id")
        send_file = bool(tcfg.get("send_file", True))
        send_only = bool(tcfg.get("send_only_if_breakouts", True))

        if not bot_token or not chat_id:
            print("ERROR: bot_token or chat_id missing in config.")
            return 2

        has_breakouts = False
        try:
            has_breakouts = df['signal'].any()
        except Exception:
            has_breakouts = True if not df.empty else False

        if send_only and not has_breakouts:
            print("No breakouts and send_only_if_breakouts=True — skipping send.")
            return 0

        breakers = df[df['signal']==True]['ticker'].tolist() if 'signal' in df.columns else []

        if breakers:
            text = f"<b>GLB Breakouts ({now})</b>\n" + "\n".join(f"• {t}" for t in breakers)
        else:
            text = f"<b>GLB Report ({now})</b>\nNo breakouts found."

        print("Sending Telegram message...")
        ok, resp = send_telegram_message(bot_token, chat_id, text)
        print("Message sent:", ok, str(resp)[:400])

        if send_file and os.path.exists(out_name):
            print("Sending Excel file...")
            ok2, resp2 = send_telegram_file(bot_token, chat_id, out_name, caption="GLB daily report")
            print("File sent:", ok2, str(resp2)[:400])

        return 0
    except Exception as e:
        print("Unhandled exception in main():", repr(e))
        traceback.print_exc()
        return 3

if __name__ == '__main__':
    sys.exit(main())
